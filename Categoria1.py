import random
import Decisiones1
from openpyxl import load_workbook
import pandas as pd
import numpy as np


def Pregunta(Nombre, Premio):
    
    wb = load_workbook('Preguntas.xlsx')
    hoja = wb.active
    ws1=wb['Hoja1']
    i=1
    Preguntas=[]
    NumeroEtiqueta=[]
    maxs=ws1.max_row
    while(i<=maxs):
        if(hoja['B%d' % i].value=='1'):
           Preguntas.append(hoja['A%d' % i].value)
           NumeroEtiqueta.append(hoja['C%d' % i].value)
        i+=1
    
    Num = random.randint(0,len(Preguntas)-1)
    print(Preguntas[Num])
    print('Digita la respuesta que consideres correcta segun las siguientes opciones:')
    wb = load_workbook('Respuestas.xlsx')
    hoja1 = wb.active
    ws2=wb['Hoja1']
    Respuestas=[]
    Correcta=[]
    i=1
    maxs2=ws2.max_row
    while(i<=maxs2):
        if(hoja1['B%d' % i].value==NumeroEtiqueta[Num]):
            Respuestas.append(hoja1['A%d' % i].value)
            Correcta.append(hoja1['C%d' % i].value)
        i+=1
    print(Respuestas)

    print('')
    print('Antes de responder deseas salirte? -->  Si (s), No (n)')
    print('')
    Decision=input()
    #Premio=100
    Premio, Decision = Decisiones1.Decisiones_1(Decision,Correcta,Nombre,Respuestas,Premio)
    if (Decision == 's'):
        print('Ganaste 0 pesos')
    else:
        Premio = 100
        print('Ganaste '+str(Premio) +' pesos')
        return (Premio,Decision)
   

