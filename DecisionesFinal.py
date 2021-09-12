from openpyxl import load_workbook
def Decisiones_Fin(Decision,Correcta,Nombre,Respuestas,Premio):
    if (Decision == 's'):
        print('Fin')
        wb3 = load_workbook('Jugadores.xlsx')
        hoja2 = wb3.active
        ws3 = wb3['Hoja1']
        maxs3 = ws3.max_row
        hoja2['A%d' % (maxs3 + 1)] = Nombre
        hoja2['B%d' % (maxs3 + 1)] = Premio
        wb3.save('Jugadores.xlsx')
        return (Premio, Decision)
    else:
        print('Cual es tu respuesta')
        Respondio = input();
        i = 0;
        while (i < len(Respuestas)):
            RespuestaS = Respuestas[i];
            if (RespuestaS == Respondio):
                NumeroRespuesta = i;
            i += 1;
        if (Correcta[NumeroRespuesta] == '1'):
            Premio=Premio+5000
            print('Ganaste 5000 pesos!')
            print('Has ganado!!! :D')
            print('Tu premio acumulado es: 6800 pesos!')
            print('')
            print('Gracias por jugar Preguntas y Respuestas! UwU')
            print('')

            wb3 = load_workbook('Jugadores.xlsx')
            hoja2 = wb3.active
            ws3 = wb3['Hoja1']
            maxs3 = ws3.max_row
            hoja2['A%d' % (maxs3 + 1)] = Nombre
            hoja2['B%d' % (maxs3 + 1)] = Premio
            wb3.save('Jugadores.xlsx')
            Premio = 0
            print('Has ganado :D ')
            return (Premio, Decision)
        else:
            wb3 = load_workbook('Jugadores.xlsx')
            hoja2 = wb3.active
            ws3 = wb3['Hoja1']
            Premio = 0
            maxs3 = ws3.max_row
            hoja2['A%d' % (maxs3 + 1)] = Nombre
            hoja2['B%d' % (maxs3 + 1)] = Premio
            wb3.save('Jugadores.xlsx')
    return (Premio, Decision)