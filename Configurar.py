#Pregunta y Respuesta Entrada - Salida
from openpyxl import load_workbook
import pandas as pd
import numpy as np

def Configuracion():
    print('Escribe el Nivel o Categoria (1,2,3,4,5)')
    Nivel=input();
    print('Escribe la pregunta')
    Pregunta=input()
    print('Escribe las respuestas a la pregunta: (1) Correcta y (3) Incorrectas')
    Respuesta1=input()
    Respuesta2=input()
    Respuesta3=input()
    Respuesta4=input()
    print('Escribe la respuesta correcta:')
    RespuestaCorrecta=input();
    Respuestas=[Respuesta1, Respuesta2, Respuesta3, Respuesta4];

    i=0;
    while(i<len(Respuestas)):
        RespuestaS=Respuestas[i];
        if(RespuestaS==RespuestaCorrecta):
            NumeroRespuesta=i;
        i+=1;
    wb = load_workbook('Preguntas.xlsx')
    hoja = wb.active
    ws1=wb['Hoja1']
    maxs=ws1.max_row
    wb = load_workbook('Preguntas.xlsx')
    hoja = wb.active
    ws1=wb['Hoja1']
    hoja['A%d' % (maxs+1)]=Pregunta
    hoja['B%d' % (maxs+1)]=Nivel
    hoja['C%d' % (maxs+1)]=maxs
    wb.save('Preguntas.xlsx')
    wb = load_workbook('Respuestas.xlsx')
    hoja1 = wb.active
    ws1=wb['Hoja1']
    hoja1['A%d' % (maxs*4+1)]=Respuesta1
    hoja1['A%d' % (maxs*4+2)]=Respuesta2
    hoja1['A%d' % (maxs*4+3)]=Respuesta3
    hoja1['A%d' % (maxs*4+4)]=Respuesta4
    hoja1['B%d' % (maxs*4+1)]=(maxs)
    hoja1['B%d' % (maxs*4+2)]=(maxs)
    hoja1['B%d' % (maxs*4+3)]=(maxs)
    hoja1['B%d' % (maxs*4+4)]=(maxs)

    if(NumeroRespuesta==0):
        hoja1['C%d' % (maxs*4+1)]='1'
        hoja1['C%d' % (maxs*4+2)]='0'
        hoja1['C%d' % (maxs*4+3)]='0'
        hoja1['C%d' % (maxs*4+4)]='0'
    if(NumeroRespuesta==1):
        hoja1['C%d' % (maxs*4+1)]='0'
        hoja1['C%d' % (maxs*4+2)]='1'
        hoja1['C%d' % (maxs*4+3)]='0'
        hoja1['C%d' % (maxs*4+4)]='0'
    if(NumeroRespuesta==2):
        hoja1['C%d' % (maxs*4+1)]='0'
        hoja1['C%d' % (maxs*4+2)]='0'
        hoja1['C%d' % (maxs*4+3)]='1'
        hoja1['C%d' % (maxs*4+4)]='0'
    if(NumeroRespuesta==3):
        hoja1['C%d' % (maxs*4+1)]='0'
        hoja1['C%d' % (maxs*4+2)]='0'
        hoja1['C%d' % (maxs*4+3)]='0'
        hoja1['C%d' % (maxs*4+4)]='1'
    wb.save('Respuestas.xlsx')

